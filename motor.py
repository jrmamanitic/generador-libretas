# ==============================================================
# motor.py — Generador automático de libretas (por lotes)
# Sin IA, sin API keys: 100% determinista, lee Excel -> genera PDF.
# ==============================================================
import io
import re
import subprocess
import unicodedata
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).parent
RUTA_PLANTILLA_SECUNDARIA = BASE_DIR / "assets" / "PLANTILLA_SECUNDARIA.xlsx"
RUTA_PLANTILLA_PRIMARIA = BASE_DIR / "assets" / "PLANTILLA_PRIMARIA.xlsx"

_PAT_IDX = re.compile(r"!\$([A-Z]{1,2})\$1(?!:)")  # ya no depende del nombre de la hoja (CONSOLIDADO/PRIMARIA/etc.)
_PAT_GRADO = re.compile(r"([1-5])\s*(?:RO|DO|ER|TO|ERO|°|º)?\s*A[ÑN]O", re.I)
_PAT_BIM = re.compile(r"\b(I{1,3}|IV)\s*BIM", re.I)

NOMBRE_HOJA_LIBRETA = "Libreta"


def _nfc(s):
    return unicodedata.normalize("NFC", str(s))


def _normalizar_nombre_hoja(s):
    """Quita tildes, espacios y diferencias de mayúsculas para comparar
    nombres de pestaña de forma tolerante (ej. 'Libreta ', 'LIBRETA', 'Líbreta')."""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return s.strip().lower()


def _hoja_libreta(wb):
    """Devuelve la pestaña de la libreta dentro de la plantilla, tolerando
    pequeños cambios en su nombre (mayúsculas, espacios, tildes) que suelen
    ocurrir al editar la plantilla en Excel. Si no la encuentra y la
    plantilla tiene una sola pestaña, asume que esa es. Si no puede
    resolverlo, lanza un error explicando exactamente qué revisar."""
    objetivo = _normalizar_nombre_hoja(NOMBRE_HOJA_LIBRETA)
    for nombre in wb.sheetnames:
        if _normalizar_nombre_hoja(nombre) == objetivo:
            return wb[nombre]
    if len(wb.sheetnames) == 1:
        return wb[wb.sheetnames[0]]
    raise KeyError(
        f"La plantilla no tiene ninguna pestaña llamada '{NOMBRE_HOJA_LIBRETA}' "
        f"(las pestañas que tiene son: {', '.join(wb.sheetnames)}). "
        f"Verifica que la pestaña con el diseño de la libreta se llame exactamente "
        f"'{NOMBRE_HOJA_LIBRETA}' (clic derecho sobre la pestaña en Excel → Cambiar nombre)."
    )


def _soffice_cmd():
    import shutil
    exe = shutil.which("soffice") or shutil.which("soffice.exe")
    if exe:
        return exe
    for c in [r"C:\Program Files\LibreOffice\program\soffice.exe",
              r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
              "/Applications/LibreOffice.app/Contents/MacOS/soffice",
              "/usr/bin/soffice", "/usr/local/bin/soffice", "/snap/bin/libreoffice"]:
        if Path(c).exists():
            return c
    raise FileNotFoundError(
        "No se encontró LibreOffice instalado. Descárgalo de https://www.libreoffice.org/download/"
    )


def _buscar_valor_por_etiqueta(ws, etiquetas, filas=range(1, 15), cols=range(1, 10)):
    """Busca una celda cuyo texto contenga alguna de las 'etiquetas' (ej. 'GRADO') y
    devuelve el primer valor no vacío a su derecha, en la misma fila."""
    for r in filas:
        for c in cols:
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and any(e in v.upper() for e in etiquetas):
                for c2 in range(c + 1, c + 8):
                    v2 = ws.cell(r, c2).value
                    if v2 not in (None, ""):
                        return str(v2).strip()
    return None


def detectar_nivel_grado_bimestre(reg_bytes, nombre_archivo=""):
    """Detecta NIVEL (primaria/secundaria), grado y bimestre leyendo el CONTENIDO
    real del archivo (no el nombre, que puede venir mal escrito o incompleto)."""
    wb = openpyxl.load_workbook(io.BytesIO(reg_bytes), data_only=True, read_only=True)
    nivel = "PRIMARIA" if "PERS" in wb.sheetnames else "SECUNDARIA"

    grado, bimestre = None, None
    if "ASISTEN" in wb.sheetnames:
        ws = wb["ASISTEN"]
        grado = _buscar_valor_por_etiqueta(ws, ["GRADO", "AÑO"])
        bimestre = _buscar_valor_por_etiqueta(ws, ["PERIODO", "BIMESTRE"])

    if not grado:
        nfc = _nfc(nombre_archivo)
        mg = _PAT_GRADO.search(nfc)
        sufijo = "GRADO" if nivel == "PRIMARIA" else "AÑO"
        grado = f"{mg.group(1)}° {sufijo}" if mg else "GRADO/AÑO DESCONOCIDO"
    if not bimestre:
        nfc = _nfc(nombre_archivo)
        mb = _PAT_BIM.search(nfc)
        bimestre = (mb.group(1).upper() + " BIM") if mb else "BIM"

    return nivel, grado, bimestre


def detectar_grado_bimestre(nombre_archivo):
    """(Obsoleto, se mantiene por compatibilidad) Extrae grado/bimestre solo del nombre."""
    nfc = _nfc(nombre_archivo)
    mg = _PAT_GRADO.search(nfc)
    mb = _PAT_BIM.search(nfc)
    grado = f"{mg.group(1)}° AÑO" if mg else "GRADO"
    bim = (mb.group(1).upper() + " BIM") if mb else "BIM"
    return grado, bim


def leer_alumnos(reg_bytes):
    """Devuelve {n_lista: nombre} desde la hoja CONSOLIDADO del registro."""
    wb = openpyxl.load_workbook(io.BytesIO(reg_bytes), data_only=True)
    cons = wb["CONSOLIDADO"]
    alumnos = {}
    for r in range(5, cons.max_row + 1):
        n, nom = cons.cell(r, 1).value, cons.cell(r, 2).value
        if n and nom and str(nom) != "0":
            try:
                alumnos[int(float(n))] = str(nom).strip()
            except (ValueError, TypeError):
                pass
    return alumnos


def _cargar_consolidado(reg_bytes):
    """Carga la hoja CONSOLIDADO del registro UNA sola vez (se reutiliza
    para todos los alumnos del lote en vez de releer el Excel por cada uno)."""
    return openpyxl.load_workbook(io.BytesIO(reg_bytes), data_only=True)["CONSOLIDADO"]


def _leer_tutor(reg_bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(reg_bytes), data_only=True, read_only=True)
        if "ASISTEN" not in wb.sheetnames:
            return None
        ws = wb["ASISTEN"]
        return _buscar_valor_por_etiqueta(ws, ["TUTOR"])
    except Exception:
        return None


def _preparar_libreta_xlsx(reg_cons, tutor, n_lista, grado, bimestre, plantilla_bytes, out_dir: Path):
    """Prepara el .xlsx de UN alumno (llena datos, resuelve fórmulas y deja
    la hoja lista para exportar). NO convierte a PDF todavía: la conversión
    se hace en lote para todos los alumnos juntos (ver _convertir_lote_a_pdf),
    porque abrir LibreOffice una vez por alumno es lo que satura el CPU."""
    fila = n_lista + 4
    nombre = reg_cons.cell(fila, 2).value
    if not nombre or str(nombre) == "0":
        raise ValueError(f"No existe el alumno N° {n_lista}")
    nombre = str(nombre).strip()

    wb_out = openpyxl.load_workbook(io.BytesIO(plantilla_bytes), data_only=False)
    lib = _hoja_libreta(wb_out)
    lib["C6"] = n_lista
    lib["D5"] = grado
    lib["C8"] = tutor or "-"
    lib["E7"] = bimestre.replace("BIM", "BIMESTRE") if "BIM" in bimestre else bimestre

    for row in lib.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                refs = _PAT_IDX.findall(cell.value)
                if not refs:
                    continue
                idx = reg_cons[refs[-1] + "1"].value
                val = reg_cons.cell(fila, int(idx)).value if idx else None
                cell.value = val if val not in (None, "", "0") else "-"

    for nombre_hoja in list(wb_out.sheetnames):
        if wb_out[nombre_hoja].title != lib.title:
            del wb_out[nombre_hoja]

    lib.sheet_properties.pageSetUpPr.fitToPage = True
    lib.page_setup.fitToWidth = 1
    lib.page_setup.fitToHeight = 2  # fuerza que TODO quepa en 2 páginas (ancho 1 x alto 2),
                                     # se autoajusta sin depender de un % fijo ni de la fuente instalada
    lib.page_setup.scale = None

    nsafe = re.sub(r"\W+", "_", nombre)[:40]
    tmp_xlsx = out_dir / f"{n_lista:02d}_{nsafe}.xlsx"
    wb_out.save(tmp_xlsx)
    return str(tmp_xlsx), nombre


def _convertir_lote_a_pdf(rutas_xlsx, out_dir: Path, tamano_lote=30):
    """Convierte varios .xlsx a PDF abriendo LibreOffice el MENOR número de
    veces posible (una sola instancia procesa todo el lote de una vez), en
    vez de un proceso de LibreOffice por cada alumno. Esto es clave para no
    saturar el CPU en hosting gratuito (Streamlit Community Cloud, etc.).

    Si por algún motivo algún archivo del lote no genera su PDF, se
    reintenta ese archivo de forma individual para no perder solo esa
    libreta."""
    if not rutas_xlsx:
        return
    soffice = _soffice_cmd()
    rutas = [Path(p) for p in rutas_xlsx]

    for i in range(0, len(rutas), tamano_lote):
        lote = rutas[i:i + tamano_lote]
        try:
            subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(out_dir)]
                + [str(p) for p in lote],
                check=False, capture_output=True, timeout=max(180, 20 * len(lote)),
            )
        except subprocess.TimeoutExpired:
            pass

        faltantes = [p for p in lote if not p.with_suffix(".pdf").exists()]
        for p in faltantes:
            try:
                subprocess.run(
                    [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(out_dir), str(p)],
                    check=False, capture_output=True, timeout=90,
                )
            except subprocess.TimeoutExpired:
                pass


def generar_una_libreta(reg_bytes, n_lista, grado, bimestre, plantilla_bytes, out_dir: Path):
    """Genera el PDF de UN solo alumno (uso individual / compatibilidad).
    Para lotes completos usa generar_todas(), que convierte todos los PDF
    de una sola vez y es mucho más liviano en CPU."""
    reg_cons = _cargar_consolidado(reg_bytes)
    tutor = _leer_tutor(reg_bytes)
    xlsx_path, nombre = _preparar_libreta_xlsx(reg_cons, tutor, n_lista, grado, bimestre, plantilla_bytes, out_dir)
    _convertir_lote_a_pdf([xlsx_path], out_dir)
    pdf_path = Path(xlsx_path).with_suffix(".pdf")
    Path(xlsx_path).unlink(missing_ok=True)
    if not pdf_path.exists():
        raise RuntimeError("LibreOffice no generó el PDF.")
    return str(pdf_path), nombre


def validar_plantilla(plantilla_bytes):
    """Verifica, apenas se sube un archivo, que sea una plantilla utilizable
    (que se pueda abrir como Excel y que tenga una pestaña de libreta
    reconocible). Devuelve (ok: bool, mensaje: str) para mostrar de inmediato
    en la interfaz, sin esperar a que falle la generación de PDFs."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(plantilla_bytes), read_only=True)
    except Exception as e:
        return False, f"⚠️ No se pudo abrir el archivo como Excel (.xlsx): {e}"
    try:
        hoja = _hoja_libreta(wb)
    except KeyError as e:
        return False, str(e)
    return True, f"✅ Plantilla válida (usa la pestaña '{hoja.title}')."


def generar_todas(reg_bytes, nombre_archivo_reg, out_dir: Path, *,
                   plantilla_primaria_bytes=None, plantilla_secundaria_bytes=None,
                   progreso_callback=None):
    """Genera TODAS las libretas de un archivo REG. Detecta automáticamente si
    el registro es de Primaria o Secundaria según el contenido real del
    archivo, y usa SIEMPRE la plantilla que corresponda a ese nivel:
    la que se haya subido como reemplazo para ese nivel (plantilla_primaria_bytes /
    plantilla_secundaria_bytes) o, si no se subió ninguna, la incluida por
    defecto en assets/. Cada nivel es independiente: reemplazar la plantilla
    de un nivel nunca afecta al otro.

    Proceso en dos fases para no saturar el CPU:
      1) Prepara el .xlsx de cada alumno (rápido, sin LibreOffice).
      2) Convierte TODOS los .xlsx a PDF en el mínimo de llamadas a LibreOffice.
    """
    nivel, grado, bimestre = detectar_nivel_grado_bimestre(reg_bytes, nombre_archivo_reg)

    if nivel == "PRIMARIA":
        plantilla_bytes = plantilla_primaria_bytes
        if plantilla_bytes is None and RUTA_PLANTILLA_PRIMARIA.exists():
            plantilla_bytes = RUTA_PLANTILLA_PRIMARIA.read_bytes()
    else:
        plantilla_bytes = plantilla_secundaria_bytes
        if plantilla_bytes is None and RUTA_PLANTILLA_SECUNDARIA.exists():
            plantilla_bytes = RUTA_PLANTILLA_SECUNDARIA.read_bytes()

    if plantilla_bytes is None:
        raise FileNotFoundError(
            f"No hay plantilla disponible para el nivel {nivel} "
            f"(archivo: {nombre_archivo_reg}). Sube una plantilla de {nivel.title()} "
            f"en la sección 'Plantillas' de arriba."
        )

    alumnos = leer_alumnos(reg_bytes)
    reg_cons = _cargar_consolidado(reg_bytes)
    tutor = _leer_tutor(reg_bytes)

    total = len(alumnos)
    resultados = {}
    pendientes = []  # (n_lista, nombre, ruta_xlsx) — a la espera de convertirse a PDF en lote

    # --- Fase 1: preparar cada libreta (sin abrir LibreOffice) ---
    for i, n_lista in enumerate(sorted(alumnos), start=1):
        nombre_esperado = alumnos[n_lista]
        try:
            xlsx_path, nombre = _preparar_libreta_xlsx(
                reg_cons, tutor, n_lista, grado, bimestre, plantilla_bytes, out_dir
            )
            pendientes.append((n_lista, nombre, xlsx_path))
        except Exception as e:
            resultados[n_lista] = {"n": n_lista, "nombre": nombre_esperado, "ok": False, "ruta": None, "error": str(e)}
        if progreso_callback:
            progreso_callback(i, total, nombre_esperado)

    # --- Fase 2: convertir TODO el lote a PDF de una sola vez ---
    if pendientes:
        if progreso_callback:
            progreso_callback(total, total, "Convirtiendo a PDF…")
        _convertir_lote_a_pdf([p for _, _, p in pendientes], out_dir)

    for n_lista, nombre, xlsx_path in pendientes:
        xlsx_path = Path(xlsx_path)
        pdf_path = xlsx_path.with_suffix(".pdf")
        xlsx_path.unlink(missing_ok=True)
        if pdf_path.exists():
            resultados[n_lista] = {"n": n_lista, "nombre": nombre, "ok": True, "ruta": str(pdf_path), "error": None}
        else:
            resultados[n_lista] = {"n": n_lista, "nombre": nombre, "ok": False, "ruta": None,
                                    "error": "LibreOffice no generó el PDF."}

    resultados_ordenados = [resultados[n] for n in sorted(resultados)]
    return nivel, grado, bimestre, resultados_ordenados
